#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
Master Script - ניתוח דיונים בפרלמנט הבריטי
UK Parliament Debates Analysis - Complete Pipeline
=============================================================================

תרגיל 1 - עיבוד שפה טבעית
NLP Assignment 1

סקריפט זה מבצע את כל שלבי התרגיל:
1. הורדת קבצי XML
2. ניקוי טקסט והפרדת סימני פיסוק
3. למטיזציה (Lemmatization)
4. בניית מטריצות TF-IDF (Word + Lemma)
5. בניית מטריצות Word2Vec/GloVe (Word + Lemma)
6. בניית מטריצות SimCSE (Original)
7. בניית מטריצות SBERT (Original)
8. חישוב חשיבות מאפיינים (Information Gain + Chi-Square)
9. יצירת קובץ Excel עם כל התוצאות

Author: [שם שלך]
Date: November 2025
=============================================================================
"""

import os
import sys
import time
import pickle
import json
import re
import warnings
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple
import logging

# Data processing
import numpy as np
import pandas as pd
from scipy import sparse
from scipy.sparse import csr_matrix, save_npz, load_npz

# Text processing
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem import WordNetLemmatizer
import spacy

# ML & NLP
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.feature_selection import mutual_info_classif, chi2
from sklearn.preprocessing import LabelEncoder

# Word embeddings
from gensim.models import Word2Vec
import gensim.downloader as gensim_api

# Transformers
from sentence_transformers import SentenceTransformer
from transformers import AutoTokenizer, AutoModel
import torch

# XML parsing
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET

# Excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

warnings.filterwarnings('ignore')

# ===========================================================================
# הגדרות גלובליות / Global Settings
# ===========================================================================

class Config:
    """הגדרות התצורה לפרויקט"""
    
    # תיקיות
    RAW_DATA_DIR = "debates_xml"           # קבצי XML מקוריים
    CLEANED_DIR = "cleaned_texts"          # טקסטים נקיים
    LEMMA_DIR = "lemmatized_texts"         # טקסטים מלומטים
    MATRICES_DIR = "matrices"              # מטריצות
    OUTPUT_DIR = "outputs"                 # פלטים סופיים
    
    # פרמטרים
    MIN_WORD_FREQ = 5                      # תדירות מינימלית למילה
    MAX_FEATURES = 10000                   # מקסימום מאפיינים ב-TF-IDF
    EMBEDDING_DIM = 100                    # גודל וקטור embedding
    TOP_N_FEATURES = 20                    # כמה מאפיינים להציג בדו"ח
    
    # מודלים
    WORD2VEC_MODEL = "glove-wiki-gigaword-100"  # או Word2Vec מאומן
    SBERT_MODEL = "all-MiniLM-L6-v2"
    SIMCSE_MODEL = "princeton-nlp/sup-simcse-bert-base-uncased"
    
    # שפה
    LANGUAGE = "english"
    SPACY_MODEL = "en_core_web_sm"

config = Config()

# ===========================================================================
# הגדרת Logging
# ===========================================================================

def setup_logging():
    """הגדרת מערכת logging"""
    log_format = '%(asctime)s - %(levelname)s - %(message)s'
    logging.basicConfig(
        level=logging.INFO,
        format=log_format,
        handlers=[
            logging.FileHandler('parliament_analysis.log', encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()

# ===========================================================================
# שלב 1: הורדת קבצים (כבר יש לך סקריפט נפרד)
# ===========================================================================

def check_downloaded_files():
    """בדיקה אם הקבצים הורדו"""
    if not os.path.exists(config.RAW_DATA_DIR):
        logger.error(f"❌ תיקייה {config.RAW_DATA_DIR} לא קיימת!")
        logger.info("💡 הרץ קודם: python download_debates.py")
        return False
    
    xml_files = list(Path(config.RAW_DATA_DIR).glob("*.xml"))
    logger.info(f"✅ נמצאו {len(xml_files)} קבצי XML")
    
    if len(xml_files) < 100:
        logger.warning(f"⚠️  נמצאו רק {len(xml_files)} קבצים. האם ההורדה הושלמה?")
    
    return len(xml_files) > 0

# ===========================================================================
# שלב 2: ניקוי טקסט והפרדת סימני פיסוק
# ===========================================================================

class TextCleaner:
    """מנקה טקסט ומפריד סימני פיסוק"""
    
    def __init__(self):
        self.punctuation_pattern = re.compile(r'([.,!?;:\'"(){}[\]<>…—–-])')
        
    def extract_text_from_xml(self, xml_path: str) -> str:
        """חילוץ טקסט מקובץ XML"""
        try:
            with open(xml_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # ניסיון לנתח כ-XML
            try:
                soup = BeautifulSoup(content, 'lxml-xml')
                # חילוץ כל הטקסט
                text = soup.get_text(separator=' ', strip=True)
            except:
                # אם נכשל, פשוט להסיר תגיות
                text = re.sub(r'<[^>]+>', ' ', content)
            
            # ניקוי רווחים מיותרים
            text = re.sub(r'\s+', ' ', text).strip()
            return text
            
        except Exception as e:
            logger.error(f"שגיאה בקריאת {xml_path}: {e}")
            return ""
    
    def separate_punctuation(self, text: str) -> str:
        """הפרדת סימני פיסוק מהמילים"""
        # הוספת רווח לפני ואחרי כל סימן פיסוק
        text = self.punctuation_pattern.sub(r' \1 ', text)
        # ניקוי רווחים מרובים
        text = re.sub(r'\s+', ' ', text).strip()
        return text
    
    def clean_text(self, text: str) -> str:
        """ניקוי כללי של הטקסט"""
        # המרה לאותיות קטנות
        text = text.lower()
        # הפרדת סימני פיסוק
        text = self.separate_punctuation(text)
        return text
    
    def process_file(self, xml_path: str, output_path: str) -> bool:
        """עיבוד קובץ בודד"""
        try:
            # חילוץ טקסט
            text = self.extract_text_from_xml(xml_path)
            if not text:
                return False
            
            # ניקוי
            cleaned = self.clean_text(text)
            
            # שמירה
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(cleaned)
            
            return True
            
        except Exception as e:
            logger.error(f"שגיאה בעיבוד {xml_path}: {e}")
            return False
    
    def process_all_files(self):
        """עיבוד כל הקבצים"""
        logger.info("🧹 מתחיל ניקוי טקסטים...")
        
        # יצירת תיקיית פלט
        Path(config.CLEANED_DIR).mkdir(parents=True, exist_ok=True)
        
        # קבלת רשימת קבצים
        xml_files = sorted(Path(config.RAW_DATA_DIR).glob("*.xml"))
        
        successful = 0
        failed = 0
        
        for i, xml_path in enumerate(xml_files, 1):
            output_name = xml_path.stem + "_cleaned.txt"
            output_path = Path(config.CLEANED_DIR) / output_name
            
            # דילוג על קבצים קיימים
            if output_path.exists():
                successful += 1
                if i % 100 == 0:
                    logger.info(f"  [{i}/{len(xml_files)}] {output_name} - קיים, מדלג")
                continue
            
            if self.process_file(str(xml_path), str(output_path)):
                successful += 1
            else:
                failed += 1
            
            if i % 100 == 0:
                logger.info(f"  [{i}/{len(xml_files)}] עובד... (הצליחו: {successful}, נכשלו: {failed})")
        
        logger.info(f"✅ ניקוי הושלם: {successful} הצליחו, {failed} נכשלו")
        return successful > 0

# ===========================================================================
# שלב 3: למטיזציה (Lemmatization)
# ===========================================================================

class Lemmatizer:
    """מבצע למטיזציה על טקסטים"""
    
    def __init__(self):
        logger.info("📚 טוען מודל spaCy...")
        try:
            self.nlp = spacy.load(config.SPACY_MODEL)
        except:
            logger.info("מוריד מודל spaCy...")
            os.system(f"python -m spacy download {config.SPACY_MODEL}")
            self.nlp = spacy.load(config.SPACY_MODEL)
        
        # להגביל את גודל הטקסט שspaCy מעבד
        self.nlp.max_length = 2000000
    
    def lemmatize_text(self, text: str) -> str:
        """ביצוע למטיזציה על טקסט"""
        doc = self.nlp(text)
        lemmas = [token.lemma_ for token in doc]
        return ' '.join(lemmas)
    
    def process_file(self, input_path: str, output_path: str) -> bool:
        """עיבוד קובץ בודד"""
        try:
            with open(input_path, 'r', encoding='utf-8') as f:
                text = f.read()
            
            # למטיזציה
            lemmatized = self.lemmatize_text(text)
            
            # שמירה
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(lemmatized)
            
            return True
            
        except Exception as e:
            logger.error(f"שגיאה בלמטיזציה של {input_path}: {e}")
            return False
    
    def process_all_files(self):
        """עיבוד כל הקבצים"""
        logger.info("🔤 מתחיל למטיזציה...")
        
        # יצירת תיקיית פלט
        Path(config.LEMMA_DIR).mkdir(parents=True, exist_ok=True)
        
        # קבלת רשימת קבצים
        cleaned_files = sorted(Path(config.CLEANED_DIR).glob("*_cleaned.txt"))
        
        successful = 0
        failed = 0
        
        for i, input_path in enumerate(cleaned_files, 1):
            output_name = input_path.stem.replace('_cleaned', '_lemma') + '.txt'
            output_path = Path(config.LEMMA_DIR) / output_name
            
            # דילוג על קבצים קיימים
            if output_path.exists():
                successful += 1
                if i % 50 == 0:
                    logger.info(f"  [{i}/{len(cleaned_files)}] {output_name} - קיים, מדלג")
                continue
            
            if self.process_file(str(input_path), str(output_path)):
                successful += 1
            else:
                failed += 1
            
            if i % 50 == 0:
                logger.info(f"  [{i}/{len(cleaned_files)}] עובד... (הצליחו: {successful}, נכשלו: {failed})")
        
        logger.info(f"✅ למטיזציה הושלמה: {successful} הצליחו, {failed} נכשלו")
        return successful > 0

# ===========================================================================
# שלב 4: בניית מטריצות TF-IDF
# ===========================================================================

class TFIDFBuilder:
    """בונה מטריצות TF-IDF עם BM25"""
    
    def __init__(self):
        self.stop_words = set(stopwords.words(config.LANGUAGE))
    
    def load_documents(self, directory: str) -> Tuple[List[str], List[str]]:
        """טעינת מסמכים מתיקייה"""
        files = sorted(Path(directory).glob("*.txt"))
        documents = []
        filenames = []
        
        for file_path in files:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    text = f.read()
                    if text.strip():
                        documents.append(text)
                        filenames.append(file_path.stem)
            except Exception as e:
                logger.error(f"שגיאה בקריאת {file_path}: {e}")
        
        logger.info(f"  טעון {len(documents)} מסמכים מ-{directory}")
        return documents, filenames
    
    def build_tfidf_matrix(self, documents: List[str], name: str) -> Tuple[csr_matrix, List[str], TfidfVectorizer]:
        """בניית מטריצת TF-IDF"""
        logger.info(f"  בונה מטריצת TF-IDF: {name}...")
        
        # BM25-like parameters
        vectorizer = TfidfVectorizer(
            max_features=config.MAX_FEATURES,
            min_df=config.MIN_WORD_FREQ,
            stop_words=list(self.stop_words),
            lowercase=True,
            norm='l2',
            use_idf=True,
            smooth_idf=True,
            sublinear_tf=True  # BM25-like
        )
        
        # בניית המטריצה
        tfidf_matrix = vectorizer.fit_transform(documents)
        feature_names = vectorizer.get_feature_names_out()
        
        logger.info(f"    מטריצה: {tfidf_matrix.shape} ({tfidf_matrix.nnz:,} ערכים שאינם אפס)")
        logger.info(f"    מאפיינים: {len(feature_names)}")
        
        return tfidf_matrix, feature_names, vectorizer
    
    def save_matrix(self, matrix: csr_matrix, feature_names: List[str], 
                   filenames: List[str], name: str):
        """שמירת מטריצה"""
        output_dir = Path(config.MATRICES_DIR) / name
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # שמירת המטריצה (sparse)
        save_npz(output_dir / "matrix.npz", matrix)
        
        # שמירת שמות מאפיינים
        with open(output_dir / "feature_names.txt", 'w', encoding='utf-8') as f:
            f.write('\n'.join(feature_names))
        
        # שמירת שמות קבצים
        with open(output_dir / "filenames.txt", 'w', encoding='utf-8') as f:
            f.write('\n'.join(filenames))
        
        # שמירת מטא-דאטה
        metadata = {
            'shape': matrix.shape,
            'nnz': matrix.nnz,
            'n_features': len(feature_names),
            'n_documents': len(filenames),
            'created': datetime.now().isoformat()
        }
        with open(output_dir / "metadata.json", 'w') as f:
            json.dump(metadata, f, indent=2)
        
        logger.info(f"    💾 נשמר ב-{output_dir}")
    
    def build_all_matrices(self):
        """בניית כל מטריצות ה-TF-IDF"""
        logger.info("📊 בונה מטריצות TF-IDF...")
        
        # TFIDF-Word
        docs_word, files_word = self.load_documents(config.CLEANED_DIR)
        if docs_word:
            matrix_word, features_word, vec_word = self.build_tfidf_matrix(docs_word, "TFIDF-Word")
            self.save_matrix(matrix_word, features_word, files_word, "TFIDF-Word")
        
        # TFIDF-Lemm
        docs_lemma, files_lemma = self.load_documents(config.LEMMA_DIR)
        if docs_lemma:
            matrix_lemma, features_lemma, vec_lemma = self.build_tfidf_matrix(docs_lemma, "TFIDF-Lemm")
            self.save_matrix(matrix_lemma, features_lemma, files_lemma, "TFIDF-Lemm")
        
        logger.info("✅ מטריצות TF-IDF הושלמו")
        
        return {
            'word': (matrix_word, features_word, files_word),
            'lemma': (matrix_lemma, features_lemma, files_lemma)
        }

# ===========================================================================
# שלב 5: בניית וקטורי Word2Vec/GloVe
# ===========================================================================

class Word2VecBuilder:
    """בונה וקטורי Word2Vec/GloVe למסמכים"""
    
    def __init__(self):
        self.stop_words = set(stopwords.words(config.LANGUAGE))
        self.model = None
    
    def load_pretrained_model(self):
        """טעינת מודל מאומן מראש"""
        logger.info(f"📥 טוען מודל: {config.WORD2VEC_MODEL}...")
        try:
            self.model = gensim_api.load(config.WORD2VEC_MODEL)
            logger.info(f"  ✅ מודל נטען: {len(self.model)} מילים, ממד {self.model.vector_size}")
        except Exception as e:
            logger.error(f"  ❌ שגיאה בטעינת מודל: {e}")
            logger.info("  💡 מאמן מודל חדש...")
            self.train_word2vec_model()
    
    def train_word2vec_model(self):
        """אימון מודל Word2Vec חדש"""
        # טעינת כל הטקסטים
        docs, _ = self.load_documents_for_training(config.CLEANED_DIR)
        
        # אימון
        sentences = [doc.split() for doc in docs]
        self.model = Word2Vec(
            sentences=sentences,
            vector_size=config.EMBEDDING_DIM,
            window=5,
            min_count=config.MIN_WORD_FREQ,
            workers=4,
            epochs=10
        )
        
        logger.info(f"  ✅ מודל אומן: {len(self.model.wv)} מילים")
    
    def load_documents_for_training(self, directory: str) -> Tuple[List[str], List[str]]:
        """טעינת מסמכים מתיקייה (ללא פיסוק ומספרים)"""
        files = sorted(Path(directory).glob("*.txt"))
        documents = []
        filenames = []
        
        for file_path in files:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    text = f.read()
                    # הסרת סימני פיסוק, מספרים, stopwords
                    text = re.sub(r'[^\w\s]', '', text)  # הסרת פיסוק
                    text = re.sub(r'\d+', '', text)  # הסרת מספרים
                    words = text.split()
                    words = [w for w in words if w not in self.stop_words and len(w) > 2]
                    cleaned_text = ' '.join(words)
                    
                    if cleaned_text.strip():
                        documents.append(cleaned_text)
                        filenames.append(file_path.stem)
            except Exception as e:
                logger.error(f"שגיאה בקריאת {file_path}: {e}")
        
        logger.info(f"  טעון {len(documents)} מסמכים מ-{directory}")
        return documents, filenames
    
    def document_to_vector(self, document: str) -> np.ndarray:
        """המרת מסמך לוקטור (ממוצע של וקטורי המילים)"""
        words = document.split()
        vectors = []
        
        for word in words:
            if word in self.model.wv:
                vectors.append(self.model.wv[word])
        
        if vectors:
            return np.mean(vectors, axis=0)
        else:
            return np.zeros(self.model.vector_size)
    
    def build_document_vectors(self, documents: List[str], name: str) -> np.ndarray:
        """בניית מטריצת וקטורים למסמכים"""
        logger.info(f"  בונה וקטורים: {name}...")
        
        vectors = []
        for doc in documents:
            vec = self.document_to_vector(doc)
            vectors.append(vec)
        
        matrix = np.vstack(vectors)
        logger.info(f"    מטריצה: {matrix.shape}")
        
        return matrix
    
    def save_matrix(self, matrix: np.ndarray, filenames: List[str], name: str):
        """שמירת מטריצה"""
        output_dir = Path(config.MATRICES_DIR) / name
        output_dir.mkdir(parents=True, exist_ok=True)
        
        # שמירת המטריצה
        np.save(output_dir / "matrix.npy", matrix)
        
        # שמירת שמות קבצים
        with open(output_dir / "filenames.txt", 'w', encoding='utf-8') as f:
            f.write('\n'.join(filenames))
        
        # מטא-דאטה
        metadata = {
            'shape': matrix.shape,
            'n_documents': len(filenames),
            'embedding_dim': matrix.shape[1],
            'created': datetime.now().isoformat()
        }
        with open(output_dir / "metadata.json", 'w') as f:
            json.dump(metadata, f, indent=2)
        
        logger.info(f"    💾 נשמר ב-{output_dir}")
    
    def build_all_matrices(self):
        """בניית כל מטריצות Word2Vec/GloVe"""
        logger.info("🔤 בונה מטריצות Word2Vec/GloVe...")
        
        # טעינת מודל
        self.load_pretrained_model()
        
        if self.model is None:
            logger.error("❌ לא ניתן לטעון מודל")
            return
        
        # W2V-Word
        docs_word, files_word = self.load_documents_for_training(config.CLEANED_DIR)
        if docs_word:
            matrix_word = self.build_document_vectors(docs_word, "W2V-Word")
            self.save_matrix(matrix_word, files_word, "W2V-Word")
        
        # W2V-Lemm
        docs_lemma, files_lemma = self.load_documents_for_training(config.LEMMA_DIR)
        if docs_lemma:
            matrix_lemma = self.build_document_vectors(docs_lemma, "W2V-Lemm")
            self.save_matrix(matrix_lemma, files_lemma, "W2V-Lemm")
        
        logger.info("✅ מטריצות Word2Vec/GloVe הושלמו")

# ===========================================================================
# שלב 6: SimCSE Embeddings
# ===========================================================================

class SimCSEBuilder:
    """בונה embeddings עם SimCSE"""
    
    def __init__(self):
        logger.info(f"🤖 טוען מודל SimCSE: {config.SIMCSE_MODEL}...")
        try:
            self.tokenizer = AutoTokenizer.from_pretrained(config.SIMCSE_MODEL)
            self.model = AutoModel.from_pretrained(config.SIMCSE_MODEL)
            self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
            self.model.to(self.device)
            logger.info(f"  ✅ מודל נטען על {self.device}")
        except Exception as e:
            logger.error(f"  ❌ שגיאה בטעינת מודל: {e}")
            self.model = None
    
    def load_original_documents(self) -> Tuple[List[str], List[str]]:
        """טעינת מסמכים מקוריים"""
        xml_files = sorted(Path(config.RAW_DATA_DIR).glob("*.xml"))
        documents = []
        filenames = []
        
        cleaner = TextCleaner()
        
        for xml_path in xml_files:
            text = cleaner.extract_text_from_xml(str(xml_path))
            if text:
                documents.append(text)
                filenames.append(xml_path.stem)
        
        logger.info(f"  טעון {len(documents)} מסמכים מקוריים")
        return documents, filenames
    
    def encode_documents(self, documents: List[str], batch_size: int = 32) -> np.ndarray:
        """קידוד מסמכים לembeddings"""
        logger.info(f"  מקודד {len(documents)} מסמכים...")
        
        embeddings = []
        
        for i in range(0, len(documents), batch_size):
            batch = documents[i:i+batch_size]
            
            # Tokenization
            inputs = self.tokenizer(
                batch,
                padding=True,
                truncation=True,
                max_length=512,
                return_tensors='pt'
            ).to(self.device)
            
            # Forward pass
            with torch.no_grad():
                outputs = self.model(**inputs)
                # Mean pooling
                batch_embeddings = outputs.last_hidden_state.mean(dim=1)
                embeddings.append(batch_embeddings.cpu().numpy())
            
            if (i // batch_size + 1) % 10 == 0:
                logger.info(f"    עובד... {i+len(batch)}/{len(documents)}")
        
        matrix = np.vstack(embeddings)
        logger.info(f"    מטריצה: {matrix.shape}")
        
        return matrix
    
    def build_and_save(self):
        """בניית ושמירת embeddings"""
        if self.model is None:
            logger.error("❌ מודל לא זמין")
            return
        
        logger.info("🧠 בונה SimCSE embeddings...")
        
        # טעינת מסמכים
        documents, filenames = self.load_original_documents()
        
        if not documents:
            logger.error("❌ לא נמצאו מסמכים")
            return
        
        # קידוד
        matrix = self.encode_documents(documents)
        
        # שמירה
        output_dir = Path(config.MATRICES_DIR) / "SimCSE-Origin"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        np.save(output_dir / "matrix.npy", matrix)
        
        with open(output_dir / "filenames.txt", 'w', encoding='utf-8') as f:
            f.write('\n'.join(filenames))
        
        metadata = {
            'shape': matrix.shape,
            'n_documents': len(filenames),
            'embedding_dim': matrix.shape[1],
            'model': config.SIMCSE_MODEL,
            'created': datetime.now().isoformat()
        }
        with open(output_dir / "metadata.json", 'w') as f:
            json.dump(metadata, f, indent=2)
        
        logger.info(f"  💾 נשמר ב-{output_dir}")
        logger.info("✅ SimCSE embeddings הושלם")

# ===========================================================================
# שלב 7: SBERT Embeddings
# ===========================================================================

class SBERTBuilder:
    """בונה embeddings עם Sentence-BERT"""
    
    def __init__(self):
        logger.info(f"🤖 טוען מודל SBERT: {config.SBERT_MODEL}...")
        try:
            self.model = SentenceTransformer(config.SBERT_MODEL)
            logger.info(f"  ✅ מודל נטען")
        except Exception as e:
            logger.error(f"  ❌ שגיאה בטעינת מודל: {e}")
            self.model = None
    
    def load_original_documents(self) -> Tuple[List[str], List[str]]:
        """טעינת מסמכים מקוריים"""
        xml_files = sorted(Path(config.RAW_DATA_DIR).glob("*.xml"))
        documents = []
        filenames = []
        
        cleaner = TextCleaner()
        
        for xml_path in xml_files:
            text = cleaner.extract_text_from_xml(str(xml_path))
            if text:
                documents.append(text)
                filenames.append(xml_path.stem)
        
        logger.info(f"  טעון {len(documents)} מסמכים מקוריים")
        return documents, filenames
    
    def encode_documents(self, documents: List[str], batch_size: int = 32) -> np.ndarray:
        """קידוד מסמכים לembeddings"""
        logger.info(f"  מקודד {len(documents)} מסמכים...")
        
        embeddings = self.model.encode(
            documents,
            batch_size=batch_size,
            show_progress_bar=True,
            convert_to_numpy=True
        )
        
        logger.info(f"    מטריצה: {embeddings.shape}")
        return embeddings
    
    def build_and_save(self):
        """בניית ושמירת embeddings"""
        if self.model is None:
            logger.error("❌ מודל לא זמין")
            return
        
        logger.info("🧠 בונה SBERT embeddings...")
        
        # טעינת מסמכים
        documents, filenames = self.load_original_documents()
        
        if not documents:
            logger.error("❌ לא נמצאו מסמכים")
            return
        
        # קידוד
        matrix = self.encode_documents(documents)
        
        # שמירה
        output_dir = Path(config.MATRICES_DIR) / "SBERT-Origin"
        output_dir.mkdir(parents=True, exist_ok=True)
        
        np.save(output_dir / "matrix.npy", matrix)
        
        with open(output_dir / "filenames.txt", 'w', encoding='utf-8') as f:
            f.write('\n'.join(filenames))
        
        metadata = {
            'shape': matrix.shape,
            'n_documents': len(filenames),
            'embedding_dim': matrix.shape[1],
            'model': config.SBERT_MODEL,
            'created': datetime.now().isoformat()
        }
        with open(output_dir / "metadata.json", 'w') as f:
            json.dump(metadata, f, indent=2)
        
        logger.info(f"  💾 נשמר ב-{output_dir}")
        logger.info("✅ SBERT embeddings הושלם")

# ===========================================================================
# שלב 8: חישוב חשיבות מאפיינים
# ===========================================================================

class FeatureImportanceCalculator:
    """מחשב חשיבות מאפיינים"""
    
    def __init__(self):
        pass
    
    def create_dummy_labels(self, n_samples: int) -> np.ndarray:
        """יצירת תוויות דמה על בסיס זמן (חצי ראשון לעומת חצי שני)"""
        # פיצול לשתי קבוצות: חצי ראשון = 0, חצי שני = 1
        labels = np.zeros(n_samples, dtype=int)
        labels[n_samples//2:] = 1
        return labels
    
    def calculate_information_gain(self, X: csr_matrix, y: np.ndarray) -> np.ndarray:
        """חישוב Information Gain"""
        logger.info("    מחשב Information Gain...")
        scores = mutual_info_classif(X, y, random_state=42)
        return scores
    
    def calculate_chi_square(self, X: csr_matrix, y: np.ndarray) -> np.ndarray:
        """חישוב Chi-Square"""
        logger.info("    מחשב Chi-Square...")
        # Chi-square דורש ערכים לא שליליים
        X_positive = X.copy()
        X_positive.data = np.abs(X_positive.data)
        scores, _ = chi2(X_positive, y)
        return scores
    
    def load_tfidf_matrix(self, name: str) -> Tuple[csr_matrix, List[str]]:
        """טעינת מטריצת TF-IDF"""
        matrix_dir = Path(config.MATRICES_DIR) / name
        
        matrix = load_npz(matrix_dir / "matrix.npz")
        
        with open(matrix_dir / "feature_names.txt", 'r', encoding='utf-8') as f:
            feature_names = [line.strip() for line in f]
        
        return matrix, feature_names
    
    def calculate_for_matrix(self, name: str) -> Dict:
        """חישוב חשיבות עבור מטריצה"""
        logger.info(f"📊 מחשב חשיבות מאפיינים עבור {name}...")
        
        # טעינת מטריצה
        matrix, feature_names = self.load_tfidf_matrix(name)
        
        # יצירת תוויות דמה
        labels = self.create_dummy_labels(matrix.shape[0])
        
        # חישוב Information Gain
        ig_scores = self.calculate_information_gain(matrix, labels)
        
        # חישוב Chi-Square
        chi2_scores = self.calculate_chi_square(matrix, labels)
        
        # יצירת DataFrame
        df = pd.DataFrame({
            'feature': feature_names,
            'information_gain': ig_scores,
            'chi_square': chi2_scores
        })
        
        # מיון לפי Information Gain
        df_ig = df.sort_values('information_gain', ascending=False).reset_index(drop=True)
        
        # מיון לפי Chi-Square
        df_chi = df.sort_values('chi_square', ascending=False).reset_index(drop=True)
        
        logger.info(f"  ✅ הושלם עבור {name}")
        
        return {
            'information_gain': df_ig,
            'chi_square': df_chi
        }
    
    def calculate_all(self) -> Dict:
        """חישוב חשיבות לכל המטריצות"""
        logger.info("🎯 מחשב חשיבות מאפיינים...")
        
        results = {}
        
        # TFIDF-Word
        try:
            results['TFIDF-Word'] = self.calculate_for_matrix('TFIDF-Word')
        except Exception as e:
            logger.error(f"שגיאה ב-TFIDF-Word: {e}")
        
        # TFIDF-Lemm
        try:
            results['TFIDF-Lemm'] = self.calculate_for_matrix('TFIDF-Lemm')
        except Exception as e:
            logger.error(f"שגיאה ב-TFIDF-Lemm: {e}")
        
        logger.info("✅ חישוב חשיבות מאפיינים הושלם")
        
        return results

# ===========================================================================
# שלב 9: יצירת קובץ Excel
# ===========================================================================

class ExcelReportGenerator:
    """מייצר דו"ח Excel עם כל התוצאות"""
    
    def __init__(self, feature_importance_results: Dict):
        self.results = feature_importance_results
        self.workbook = Workbook()
        # מחיקת גיליון ברירת המחדל
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
    
    def create_sheet(self, name: str, df: pd.DataFrame, metric: str):
        """יצירת גיליון בודד"""
        ws = self.workbook.create_sheet(title=name[:31])  # Excel מגביל ל-31 תווים
        
        # כותרת
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # כתיבת כותרות
        headers = ['Rank', 'Feature', metric.replace('_', ' ').title()]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # כתיבת נתונים
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=row_idx-1)  # Rank
            ws.cell(row=row_idx, column=2, value=row['feature'])
            ws.cell(row=row_idx, column=3, value=float(row[metric]))
        
        # התאמת רוחב עמודות
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
    
    def generate(self, output_path: str):
        """יצירת הקובץ"""
        logger.info("📝 מייצר קובץ Excel...")
        
        for matrix_name, metrics in self.results.items():
            # Information Gain
            df_ig = metrics['information_gain']
            self.create_sheet(f"{matrix_name}-IG", df_ig, 'information_gain')
            
            # Chi-Square
            df_chi = metrics['chi_square']
            self.create_sheet(f"{matrix_name}-Chi2", df_chi, 'chi_square')
        
        # שמירה
        self.workbook.save(output_path)
        logger.info(f"  💾 נשמר ב-{output_path}")
        logger.info("✅ קובץ Excel הושלם")

# ===========================================================================
# שלב 10: ארגון ודחיסת קבצים
# ===========================================================================

class FilesOrganizer:
    """מארגן ודוחס את כל הקבצים להגשה"""
    
    def __init__(self, student_names: str):
        self.student_names = student_names
        self.submission_dir = Path(student_names)
    
    def organize_files(self):
        """ארגון הקבצים"""
        logger.info("📦 מארגן קבצים להגשה...")
        
        # יצירת תיקייה ראשית
        self.submission_dir.mkdir(exist_ok=True)
        
        # העתקת קוד
        import shutil
        if os.path.exists('parliament_analysis_master.py'):
            shutil.copy('parliament_analysis_master.py', self.submission_dir / 'code.py')
        
        # העתקת README (אם קיים)
        if os.path.exists('README.pdf'):
            shutil.copy('README.pdf', self.submission_dir)
        elif os.path.exists('README.docx'):
            shutil.copy('README.docx', self.submission_dir)
        
        # העתקת Excel
        excel_file = Path(config.OUTPUT_DIR) / 'feature_importance_results.xlsx'
        if excel_file.exists():
            shutil.copy(excel_file, self.submission_dir / 'features.xlsx')
        
        logger.info("  ✅ קבצים אורגנו")
    
    def zip_matrices(self):
        """דחיסת כל מטריצה בנפרד"""
        logger.info("🗜️  דוחס מטריצות...")
        
        import zipfile
        
        matrices_dir = Path(config.MATRICES_DIR)
        
        for matrix_folder in matrices_dir.iterdir():
            if matrix_folder.is_dir():
                zip_name = self.submission_dir / f"{matrix_folder.name}.zip"
                
                with zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file in matrix_folder.rglob('*'):
                        if file.is_file():
                            zipf.write(file, file.relative_to(matrices_dir))
                
                logger.info(f"  ✅ {matrix_folder.name}.zip")
        
        logger.info("  ✅ כל המטריצות נדחסו")
    
    def zip_submission(self):
        """דחיסת תיקיית ההגשה"""
        logger.info("🗜️  דוחס הגשה סופית...")
        
        import zipfile
        
        zip_name = f"{self.student_names}.zip"
        
        with zipfile.ZipFile(zip_name, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file in self.submission_dir.rglob('*'):
                if file.is_file():
                    zipf.write(file, file.relative_to(self.submission_dir.parent))
        
        logger.info(f"  💾 נוצר: {zip_name}")
        logger.info("✅ הגשה מוכנה!")

# ===========================================================================
# תזמון ראשי - Main Pipeline
# ===========================================================================

class MasterPipeline:
    """מנהל את כל התהליך"""
    
    def __init__(self):
        self.start_time = time.time()
        
        # יצירת תיקיות בסיסיות
        for directory in [config.CLEANED_DIR, config.LEMMA_DIR, 
                         config.MATRICES_DIR, config.OUTPUT_DIR]:
            Path(directory).mkdir(parents=True, exist_ok=True)
    
    def download_nltk_data(self):
        """הורדת נתוני NLTK"""
        logger.info("📚 מוריד נתוני NLTK...")
        try:
            nltk.download('stopwords', quiet=True)
            nltk.download('punkt', quiet=True)
            nltk.download('wordnet', quiet=True)
            nltk.download('averaged_perceptron_tagger', quiet=True)
            logger.info("  ✅ נתוני NLTK הורדו")
        except Exception as e:
            logger.warning(f"  ⚠️  שגיאה בהורדת NLTK: {e}")
    
    def run_step(self, step_num: int, step_name: str, func, *args, **kwargs):
        """הרצת שלב בודד עם מדידת זמן"""
        logger.info("")
        logger.info("=" * 70)
        logger.info(f"שלב {step_num}: {step_name}")
        logger.info("=" * 70)
        
        step_start = time.time()
        try:
            result = func(*args, **kwargs)
            step_time = time.time() - step_start
            logger.info(f"⏱️  זמן שלב: {step_time/60:.2f} דקות")
            return result
        except Exception as e:
            logger.error(f"❌ שגיאה בשלב {step_num}: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def run_full_pipeline(self, skip_download: bool = True):
        """הרצת כל התהליך"""
        logger.info("")
        logger.info("*" * 70)
        logger.info("  🚀 מתחיל תהליך ניתוח דיוני הפרלמנט הבריטי")
        logger.info("*" * 70)
        logger.info("")
        
        # הכנות
        self.download_nltk_data()
        
        # שלב 1: בדיקת קבצים
        self.run_step(1, "בדיקת קבצים שהורדו", check_downloaded_files)
        
        # שלב 2: ניקוי טקסט
        cleaner = TextCleaner()
        self.run_step(2, "ניקוי טקסט והפרדת סימני פיסוק", cleaner.process_all_files)
        
        # שלב 3: למטיזציה
        lemmatizer = Lemmatizer()
        self.run_step(3, "למטיזציה", lemmatizer.process_all_files)
        
        # שלב 4: TF-IDF
        tfidf_builder = TFIDFBuilder()
        self.run_step(4, "בניית מטריצות TF-IDF", tfidf_builder.build_all_matrices)
        
        # שלב 5: Word2Vec/GloVe
        w2v_builder = Word2VecBuilder()
        self.run_step(5, "בניית מטריצות Word2Vec/GloVe", w2v_builder.build_all_matrices)
        
        # שלב 6: SimCSE
        simcse_builder = SimCSEBuilder()
        self.run_step(6, "בניית embeddings SimCSE", simcse_builder.build_and_save)
        
        # שלב 7: SBERT
        sbert_builder = SBERTBuilder()
        self.run_step(7, "בניית embeddings SBERT", sbert_builder.build_and_save)
        
        # שלב 8: חישוב חשיבות מאפיינים
        importance_calc = FeatureImportanceCalculator()
        results = self.run_step(8, "חישוב חשיבות מאפיינים", importance_calc.calculate_all)
        
        # שלב 9: יצירת Excel
        if results:
            excel_gen = ExcelReportGenerator(results)
            excel_path = Path(config.OUTPUT_DIR) / 'feature_importance_results.xlsx'
            self.run_step(9, "יצירת קובץ Excel", excel_gen.generate, str(excel_path))
        
        # סיכום
        total_time = time.time() - self.start_time
        logger.info("")
        logger.info("=" * 70)
        logger.info("🎉 התהליך הושלם!")
        logger.info("=" * 70)
        logger.info(f"⏱️  זמן כולל: {total_time/60:.2f} דקות ({total_time/3600:.2f} שעות)")
        logger.info("")
        logger.info("📁 קבצים שנוצרו:")
        logger.info(f"  • טקסטים נקיים: {config.CLEANED_DIR}/")
        logger.info(f"  • טקסטים מלומטים: {config.LEMMA_DIR}/")
        logger.info(f"  • מטריצות: {config.MATRICES_DIR}/")
        logger.info(f"  • תוצאות: {config.OUTPUT_DIR}/")
        logger.info("")
        logger.info("📋 השלבים הבאים:")
        logger.info("  1. צור קובץ README עם הסברים מפורטים")
        logger.info("  2. הרץ organize_submission.py לארגון הקבצים")
        logger.info("  3. העלה את הקובץ המזופזף למודל")
        logger.info("")

# ===========================================================================
# ארגון הגשה
# ===========================================================================

def organize_submission(student_names: str = "StudentNames"):
    """ארגון קבצים להגשה"""
    organizer = FilesOrganizer(student_names)
    organizer.organize_files()
    organizer.zip_matrices()
    organizer.zip_submission()

# ===========================================================================
# Main Entry Point
# ===========================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Master Script - ניתוח דיוני הפרלמנט הבריטי',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument(
        '--skip-download',
        action='store_true',
        help='דלג על שלב ההורדה (הקבצים כבר קיימים)'
    )
    
    parser.add_argument(
        '--organize',
        action='store_true',
        help='רק ארגן קבצים להגשה'
    )
    
    parser.add_argument(
        '--student-names',
        type=str,
        default='StudentNames',
        help='שמות הסטודנטים להגשה'
    )
    
    args = parser.parse_args()
    
    try:
        if args.organize:
            organize_submission(args.student_names)
        else:
            pipeline = MasterPipeline()
            pipeline.run_full_pipeline(skip_download=args.skip_download)
    
    except KeyboardInterrupt:
        logger.info("\n\n⚠️  התהליך הופסק על ידי המשתמש")
    except Exception as e:
        logger.error(f"\n❌ שגיאה כללית: {e}")
        import traceback
        traceback.print_exc()